import sys
import pandas as pd
import os
import glob
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border, Side


def get_ob_identification_results(data_frame):
    n_ob_in_gt = data_frame['#OB Sentences in Ground Truths'].sum()
    n_ob_in_res = data_frame['#OB Sentences in Responses'].sum()
    n_ob_tp = data_frame['OB Sentences - TP'].sum()
    n_ob_fp = data_frame['OB Sentences - FP'].sum()
    n_ob_fn = data_frame['OB Sentences - FN'].sum()
    n_ob_tn = data_frame['OB Sentences - TN'].sum()
    n_ob_precision = n_ob_tp / (n_ob_fp + n_ob_tp)
    n_ob_recall = n_ob_tp / (n_ob_tp + n_ob_fn)
    n_ob_f1 = 2 * n_ob_precision * n_ob_recall / (n_ob_precision + n_ob_recall)

    return n_ob_in_gt, n_ob_in_res, n_ob_tp, n_ob_fp, n_ob_fn, n_ob_tn, n_ob_precision, n_ob_recall, n_ob_f1


def get_eb_identification_results(data_frame):
    n_eb_in_gt = data_frame['#EB Sentences in Ground Truths'].sum()
    n_eb_in_res = data_frame['#EB Sentences in Responses'].sum()
    n_eb_tp = data_frame['EB Sentences - TP'].sum()
    n_eb_fp = data_frame['EB Sentences - FP'].sum()
    n_eb_fn = data_frame['EB Sentences - FN'].sum()
    n_eb_tn = data_frame['EB Sentences - TN'].sum()
    n_eb_precision = n_eb_tp / (n_eb_fp + n_eb_tp)
    n_eb_recall = n_eb_tp / (n_eb_tp + n_eb_fn)
    n_eb_f1 = 2 * n_eb_precision * n_eb_recall / (n_eb_precision + n_eb_recall)

    return n_eb_in_gt, n_eb_in_res, n_eb_tp, n_eb_fp, n_eb_fn, n_eb_tn, n_eb_precision, n_eb_recall, n_eb_f1


def s2r_identification_results(data_frame):
    n_s2r_in_gt = data_frame['#S2R Sentences in Ground Truths'].sum()
    n_s2r_in_res = data_frame['#S2R Sentences in Responses'].sum()
    n_s2r_tp = data_frame['S2R Sentences - TP'].sum()
    n_s2r_fp = data_frame['S2R Sentences - FP'].sum()
    n_s2r_fn = data_frame['S2R Sentences - FN'].sum()
    n_s2r_tn = data_frame['S2R Sentences - TN'].sum()
    n_s2r_precision = n_s2r_tp / (n_s2r_fp + n_s2r_tp)
    n_s2r_recall = n_s2r_tp / (n_s2r_tp + n_s2r_fn)
    n_s2r_f1 = 2 * n_s2r_precision * n_s2r_recall / (n_s2r_precision + n_s2r_recall)

    return n_s2r_in_gt, n_s2r_in_res, n_s2r_tp, n_s2r_fp, n_s2r_fn, n_s2r_tn, n_s2r_precision, n_s2r_recall, n_s2r_f1


def highlight_cells(sheet_name, columns_to_highlight, max_values):
    yellow_fill = PatternFill(start_color='00FFFF00',
                              end_color='00FFFF00',
                              fill_type='solid')

    for i in range(len(columns_to_highlight)):
        for cell in writer.sheets[sheet_name][columns_to_highlight[i]]:
            if cell.value == max_values[i]:
                cell.fill = yellow_fill


def highlight_best_results_sheet_1(data_frame, sheet_name):
    columns_to_highlight = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    max_ob_p = data_frame['OB-Precision'].max().round(4)
    max_ob_r = data_frame['OB-Recall'].max().round(4)
    max_ob_f1 = data_frame['OB-F1'].max().round(4)
    max_eb_p = data_frame['EB-Precision'].max().round(4)
    max_eb_r = data_frame['EB-Recall'].max().round(4)
    max_eb_f1 = data_frame['EB-F1'].max().round(4)
    max_s2r_p = data_frame['S2R-Precision'].max().round(4)
    max_s2r_r = data_frame['S2R-Recall'].max().round(4)
    max_s2r_f1 = data_frame['S2R-F1'].max().round(4)

    max_values = [max_ob_p, max_ob_r, max_ob_f1, max_eb_p, max_eb_r, max_eb_f1, max_s2r_p, max_s2r_r, max_s2r_f1]
    highlight_cells(sheet_name, columns_to_highlight, max_values)


def highlight_best_results_sheet_2(data_frame, sheet_name):
    columns_to_highlight = ['F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'P', 'R', 'S', 'T', 'U', 'V']
    max_ob_res = data_frame['#OB-Predicted'].max()
    max_ob_tp = data_frame['OB-TP'].max()
    max_ob_tn = data_frame['OB-TN'].max()
    min_ob_fp = data_frame['OB-FP'].min()
    min_ob_fn = data_frame['OB-FN'].min()
    max_eb_res = data_frame['#EB-Predicted'].max()
    max_eb_tp = data_frame['EB-TP'].max()
    max_eb_tn = data_frame['EB-TN'].max()
    min_eb_fp = data_frame['EB-FP'].min()
    min_eb_fn = data_frame['EB-FN'].min()
    max_s2r_res = data_frame['#S2R-Predicted'].max()
    max_s2r_tp = data_frame['S2R-TP'].max()
    max_s2r_tn = data_frame['S2R-TN'].max()
    min_s2r_fp = data_frame['S2R-FP'].min()
    min_s2r_fn = data_frame['S2R-FN'].min()

    max_min_values = [max_ob_res, max_ob_tp, max_ob_tn, min_ob_fp, min_ob_fn, max_eb_res, max_eb_tp, max_eb_tn, min_eb_fp,
                  min_eb_fn, max_s2r_res, max_s2r_tp, max_s2r_tn, min_s2r_fp, min_s2r_fn]

    highlight_cells(sheet_name, columns_to_highlight, max_min_values)


def set_border(ws, staring_range, ending_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    rows = ws[staring_range:ending_range]
    for row in rows:
        for cell in row:
            cell.border = border


def modify_df_and_save_to_excel(df, writer, sheet_name, wrap_text_columns):
    df.style.set_properties(**{'text-align': 'center'}).to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.4f")
    worksheet = writer.sheets[sheet_name]

    # Wrap text in the columns
    for col in wrap_text_columns:
        for cell in writer.sheets[sheet_name][col]:
            cell.alignment = Alignment(wrap_text=True)

    # # Set the width of the column
    # for col in reduce_size_columns:
    #     worksheet.column_dimensions[col].width = 5.3
    #
    worksheet.column_dimensions['A'].width = 38

    worksheet.freeze_panes = 'A2'

    if sheet_name.endswith('1'):
        highlight_best_results_sheet_1(df, sheet_name)
        starting_range = 'A1'
        ending_range = 'K12'
        set_border(worksheet, starting_range, ending_range)

    elif sheet_name.endswith('2'):
        highlight_best_results_sheet_2(df, sheet_name)
        starting_range = 'A1'
        ending_range = 'V12'
        set_border(worksheet, starting_range, ending_range)

    writer._save()


if __name__ == "__main__":
    results_path = './calculated_metrics'
    output_file_path = './results_summary_euler_11_08_24.xlsx'

    # exp_defs = ["BEE-Model", "ChatGPT-ZS-No_Def-Comb", "ChatGPT-ZS-No_Def-Split", "ChatGPT-ZS-One_Line_Def-Comb", "ChatGPT-ZS-One_Line_Def-Split", "ChatGPT-ZS-Detailed_Def-Comb", "ChatGPT-ZS-Detailed_Def-Split", "ChatGPT-ZS-Detailed_Def_with_Ex-Comb", "ChatGPT-ZS-Detailed_Def_with_Ex-Split", "ChatGPT-FS-One_Line_Def-Comb", "ChatGPT-CoT-One_Line_Def-Comb"]
    exp_defs = ["BEE-Model", "ZS-No_Definition", "ZS-Simple_Definition", "ZS-Detailed_Definition", "FS-Simple_Definition", "CoT-Simple_Definition"]
    #exp_defs = ["ChatGPT-ZS-No_Def-Comb", "ChatGPT-ZS-One_Line_Def-Comb", "ChatGPT-ZS-Detailed_Def-Comb", "ChatGPT-FS-One_Line_Def-Comb", "ChatGPT-CoT-One_Line_Def-Comb"]

    # exp_defs = ["ChatGPT-ZS-No_Def-Comb", "ChatGPT-ZS-No_Def-Split", "ChatGPT-ZS-One_Line_Def-Comb", "ChatGPT-ZS-One_Line_Def-Split", "ChatGPT-ZS-Detailed_Def-Comb", "ChatGPT-ZS-Detailed_Def-Split", "ChatGPT-ZS-Detailed_Def_with_Ex-Comb", "ChatGPT-ZS-Detailed_Def_with_Ex-Split", "ChatGPT-FS-One_Line_Def-Comb", "ChatGPT-CoT-One_Line_Def-Comb"]

    bug_ids_not_euler = [1, 81, 104]

    # prompt_versions = [0.0, 1.0, 1.1, 1.2, 2.1, 3.1]
    wrap_text_columns_sheet1 = ['A', 'C', 'F', 'I']
    wrap_text_columns_sheet2 = ['D', 'F', 'K', 'L', 'Q', 'R']

    # Remove the output file if it exists
    if os.path.exists(output_file_path):
        os.remove(output_file_path)

    folder_paths = glob.glob(os.path.join(results_path, "*"))
    folder_paths.sort()

    # Iterate over all the dataset folders in the results folder
    for folder_path in folder_paths:
        if not os.path.isdir(folder_path):
            continue
        if "euler" in folder_path:
            #continue
            exp_defs = ["BEE-Model", "ZS-No_Definition", "ZS-Simple_Definition", "ZS-Detailed_Definition", "FS-Simple_Definition", "CoT-Simple_Definition", "EULER-Model"]


        # Get the dataset name from the folder name
        dataset_name = folder_path.split("/")[-1]
        sheet1_name = dataset_name + "-1"
        sheet2_name = dataset_name + "-2"
        prf1_rows = []
        tp_tn_fp_fn_rows = []

        subfolder_paths = glob.glob(os.path.join(folder_path, "*"))
        subfolder_paths.sort()

        folder = folder_path.split("/")[-1]
        print(folder)

        # Iterate over all the prompt version folders in the dataset folder
        for subfolder_path in subfolder_paths:
            if not os.path.isdir(subfolder_path):
                continue
            if subfolder_path.endswith("1.4"):
                continue

            # print(subfolder_path)
            # if "euler-model-0.1" in subfolder_path:
            #     continue

            file_paths = glob.glob(os.path.join(subfolder_path, "*"))
            file_paths.sort()

            # Iterate over all the files in the prompt version folder
            for file_path in file_paths:
                if not file_path.endswith(".csv"):
                    continue
                # Get the prompt version from the file name
                prompt_version = file_path.split("-")[-1].replace(".csv", "")
                if not prompt_version.endswith("0"):
                    continue

                print(file_path)
                # if not file_path.endswith("euler-metrics-0.0.0.csv"):
                #     continue
                # Read the csv file
                result_df = pd.read_csv(file_path)
                if folder == "euler-data":
                    # Drop the rows with bug ids that are not in the euler dataset
                    result_df = result_df[~result_df['Bug_id'].isin(bug_ids_not_euler)]

                n_bug_reports = len(result_df)
                print(n_bug_reports)
                n_sentences = result_df["# Sentences in Bug Reports"].sum()

                if prompt_version.endswith("0"):
                    # print(prompt_version)
                    prompt_version = prompt_version.split(".")
                    # print(prompt_version)
                    prompt_version = prompt_version[0] + "." + prompt_version[1]
                    exp_name = prompt_version
                    # print(exp_name)

                    n_ob_in_gt, n_ob_in_res, n_ob_tp, n_ob_fp, n_ob_fn, n_ob_tn, n_ob_precision, n_ob_recall, n_ob_f1 = get_ob_identification_results(
                        result_df)
                    n_eb_in_gt, n_eb_in_res, n_eb_tp, n_eb_fp, n_eb_fn, n_eb_tn, n_eb_precision, n_eb_recall, n_eb_f1 = get_eb_identification_results(
                        result_df)
                    n_s2r_in_gt, n_s2r_in_res, n_s2r_tp, n_s2r_fp, n_s2r_fn, n_s2r_tn, n_s2r_precision, n_s2r_recall, n_s2r_f1 = s2r_identification_results(
                        result_df)

                    prf1_rows.append(
                        [exp_name, n_ob_precision, n_ob_recall, n_ob_f1, n_eb_precision, n_eb_recall, n_eb_f1,
                         n_s2r_precision, n_s2r_recall, n_s2r_f1])
                    tp_tn_fp_fn_rows.append(
                        [exp_name, n_bug_reports, n_sentences, n_ob_in_gt, n_ob_in_res, n_ob_tp, n_ob_tn, n_ob_fp, n_ob_fn, n_eb_in_gt, n_eb_in_res,
                         n_eb_tp, n_eb_tn, n_eb_fp, n_eb_fn, n_s2r_in_gt, n_s2r_in_res, n_s2r_tp, n_s2r_tn, n_s2r_fp,
                         n_s2r_fn])

                else:
                    prompt_version_list = prompt_version.split(".")
                    exp_name = prompt_version_list[0] + "." + prompt_version_list[1] + "." + "x"
                    if prompt_version.endswith("1"):
                        n_ob_in_gt, n_ob_in_res, n_ob_tp, n_ob_fp, n_ob_fn, n_ob_tn, n_ob_precision, n_ob_recall, n_ob_f1 = get_ob_identification_results(
                            result_df)
                    elif prompt_version.endswith("2"):
                        n_eb_in_gt, n_eb_in_res, n_eb_tp, n_eb_fp, n_eb_fn, n_eb_tn, n_eb_precision, n_eb_recall, n_eb_f1 = get_eb_identification_results(
                            result_df)
                    elif prompt_version.endswith("3"):
                        n_s2r_in_gt, n_s2r_in_res, n_s2r_tp, n_s2r_fp, n_s2r_fn, n_s2r_tn, n_s2r_precision, n_s2r_recall, n_s2r_f1 = s2r_identification_results(
                            result_df)

            if prompt_version.endswith("0"):
                continue

            # prf1_rows.append(
            #     [exp_name, n_ob_precision, n_ob_recall, n_ob_f1, n_eb_precision, n_eb_recall, n_eb_f1, n_s2r_precision,
            #      n_s2r_recall, n_s2r_f1])
            # tp_tn_fp_fn_rows.append(
            #     [exp_name, n_bug_reports, n_sentences, n_ob_in_gt, n_ob_in_res, n_ob_tp, n_ob_tn, n_ob_fp, n_ob_fn, n_eb_in_gt, n_eb_in_res,
            #      n_eb_tp, n_eb_tn, n_eb_fp, n_eb_fn, n_s2r_in_gt, n_s2r_in_res, n_s2r_tp, n_s2r_tn, n_s2r_fp, n_s2r_fn])

        # Create the dataframe
        prf1_df = pd.DataFrame(prf1_rows,
                               columns=["Exp-Name", "OB-Precision", "OB-Recall", "OB-F1", "EB-Precision", "EB-Recall",
                                        "EB-F1", "S2R-Precision", "S2R-Recall", "S2R-F1"])
        tp_tn_fp_fn_df = pd.DataFrame(tp_tn_fp_fn_rows,
                                      columns=["Exp-Name", "Total-#BRs", "Total-#Sentences", "#OB-GT", "#OB-Predicted", "OB-TP", "OB-TN", "OB-FP",
                                               "OB-FN", "#EB-GT", "#EB-Predicted", "EB-TP", "EB-TN", "EB-FP", "EB-FN",
                                               "#S2R-GT", "#S2R-Predicted", "S2R-TP", "S2R-TN", "S2R-FP", "S2R-FN"])

        # print(exp_defs)
        #print(prf1_df)
        prf1_df.insert(0, "Exp-Definition", exp_defs)
        tp_tn_fp_fn_df.insert(0, "Exp-Definition", exp_defs)

        # Write the dataframe to the excel file
        if os.path.exists(output_file_path):
            with pd.ExcelWriter(output_file_path, mode="a", engine="openpyxl") as writer:
                modify_df_and_save_to_excel(prf1_df, writer, sheet1_name, wrap_text_columns_sheet1)
                modify_df_and_save_to_excel(tp_tn_fp_fn_df, writer, sheet2_name, wrap_text_columns_sheet2)

        else:
            with pd.ExcelWriter(output_file_path, mode="w", engine="openpyxl") as writer:
                modify_df_and_save_to_excel(prf1_df, writer, sheet1_name, wrap_text_columns_sheet1)
                modify_df_and_save_to_excel(tp_tn_fp_fn_df, writer, sheet2_name, wrap_text_columns_sheet2)
